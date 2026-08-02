"""Read a spreadsheet into validated row models, collecting every problem as it goes.

Errors carry the spreadsheet row number so you can fix the file directly rather than
reverse-engineering which record went wrong.
"""

from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from app.ingest.excel.schema import SheetSpec
from app.ingest.result import ParseResult, RowError

HEADER_ROW = 1


def _normalise(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def read_sheet(path: Path, spec: SheetSpec) -> ParseResult:
    sheet = spec.key
    result: ParseResult = ParseResult(sheet=sheet)

    if not path.exists():
        result.errors.append(RowError(sheet, None, None, f"file not found: {path}"))
        return result

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a zoo of exception types
        result.errors.append(RowError(sheet, None, None, f"could not open workbook: {exc}"))
        return result

    try:
        worksheet = workbook[spec.key] if spec.key in workbook.sheetnames else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)

        try:
            raw_header = next(rows)
        except StopIteration:
            result.errors.append(RowError(sheet, None, None, "sheet is empty"))
            return result

        header = [_normalise(c) for c in raw_header]
        position_of = {name: idx for idx, name in enumerate(header) if name}

        missing = [h for h in spec.required_headers if h not in position_of]
        if missing:
            result.errors.append(
                RowError(
                    sheet,
                    HEADER_ROW,
                    None,
                    f"missing required column(s): {', '.join(missing)}. "
                    f"Expected header row: {', '.join(spec.headers)}",
                )
            )
            return result

        unknown = [h for h in position_of if h not in spec.headers]
        if unknown:
            result.errors.append(
                RowError(
                    sheet,
                    HEADER_ROW,
                    None,
                    f"unrecognised column(s): {', '.join(sorted(unknown))}. "
                    f"Rename or remove them - the header row is the contract.",
                )
            )
            return result

        for offset, raw in enumerate(rows, start=HEADER_ROW + 1):
            if raw is None or all(_normalise(c) == "" for c in raw):
                continue  # blank spacer row

            data = {
                name: (raw[idx] if idx < len(raw) else None) for name, idx in position_of.items()
            }
            try:
                result.rows.append((offset, spec.row_model(**data)))
            except ValidationError as exc:
                for err in exc.errors():
                    field = ".".join(str(p) for p in err["loc"]) if err["loc"] else None
                    result.errors.append(RowError(sheet, offset, field, err["msg"]))
    finally:
        workbook.close()

    if not result.rows and not result.errors:
        result.errors.append(RowError(sheet, None, None, "sheet has a header but no data rows"))

    return result
