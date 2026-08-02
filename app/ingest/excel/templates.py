"""Generate blank workbooks with the correct headers, so the contract is never guesswork."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.ingest.excel.schema import ALL_SPECS, SheetSpec

_HEADER_FILL = PatternFill("solid", start_color="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_REQUIRED_FILL = PatternFill("solid", start_color="FFF2CC")


def write_template(spec: SheetSpec, directory: Path) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / spec.filename

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = spec.key

    for index, column in enumerate(spec.columns, start=1):
        cell = data_sheet.cell(row=1, column=index, value=column.name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        width = max(len(column.name), len(column.example), 12) + 4
        data_sheet.column_dimensions[get_column_letter(index)].width = width
        if column.required:
            data_sheet.cell(row=2, column=index).fill = _REQUIRED_FILL

    data_sheet.freeze_panes = "A2"

    notes = workbook.create_sheet("_notes")
    notes["A1"], notes["B1"], notes["C1"], notes["D1"] = "column", "required", "meaning", "example"
    for cell in ("A1", "B1", "C1", "D1"):
        notes[cell].fill = _HEADER_FILL
        notes[cell].font = _HEADER_FONT
    for row, column in enumerate(spec.columns, start=2):
        notes.cell(row=row, column=1, value=column.name)
        notes.cell(row=row, column=2, value="yes" if column.required else "no")
        notes.cell(row=row, column=3, value=column.help)
        notes.cell(row=row, column=4, value=column.example)
    for letter, width in (("A", 22), ("B", 10), ("C", 62), ("D", 22)):
        notes.column_dimensions[letter].width = width

    notes.cell(
        row=len(spec.columns) + 3,
        column=1,
        value=(
            "Do not rename, reorder or add columns - the header row is validated exactly. "
            "Codes are matched case-insensitively and stored upper-case. "
            "Re-importing the same file is always safe: it is an upsert, not an append."
        ),
    )

    workbook.save(path)
    return path


def write_all_templates(directory: Path) -> list[Path]:
    return [write_template(spec, directory) for spec in ALL_SPECS.values()]
